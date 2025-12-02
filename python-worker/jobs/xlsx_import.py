"""
XLSX Import Job Handler
Handles Excel file import with formula preservation
"""
import json
import hashlib
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.formula.translate import Translator
import psycopg2
from psycopg2.extras import execute_values
from utils.db import get_db_connection
from utils.s3 import download_from_s3, upload_to_s3
from utils.logger import setup_logger

logger = setup_logger(__name__)

# Volatile Excel functions that should be flagged
VOLATILE_FUNCTIONS = ['NOW', 'TODAY', 'RAND', 'RANDBETWEEN', 'RANDARRAY', 'OFFSET', 'INDIRECT', 'CELL', 'INFO']

def detect_volatile_formula(formula: str) -> bool:
    """Check if formula contains volatile functions"""
    if not formula:
        return False
    formula_upper = formula.upper()
    return any(func in formula_upper for func in VOLATILE_FUNCTIONS)

def extract_formula_references(formula: str) -> Dict[str, Any]:
    """Extract references from formula (other sheets, external workbooks)"""
    refs = {
        'external_references': [],
        'sheet_references': [],
        'volatile': False,
    }
    
    if not formula:
        return refs
    
    refs['volatile'] = detect_volatile_formula(formula)
    
    # Match sheet references like 'Sheet1'!A1 or [Workbook.xlsx]Sheet1!A1
    sheet_pattern = r"(?:\[([^\]]+)\])?([^!]+)!(?:[A-Z]+[0-9]+|[A-Z]+:[A-Z]+)"
    matches = re.findall(sheet_pattern, formula)
    
    for match in matches:
        workbook, sheet = match
        if workbook:
            refs['external_references'].append({
                'workbook': workbook,
                'sheet': sheet,
            })
        else:
            refs['sheet_references'].append(sheet)
    
    return refs

def compute_formula_hash(formula: str) -> str:
    """Compute hash of formula text for deduplication"""
    if not formula:
        return ''
    return hashlib.sha256(formula.encode('utf-8')).hexdigest()[:16]

def detect_header_row(worksheet, max_rows: int = 100) -> Optional[int]:
    """Detect header row using heuristics"""
    header_keywords = ['date', 'amount', 'description', 'category', 'currency', 'source', 'transaction']
    
    for row_idx in range(1, min(max_rows, worksheet.max_row + 1)):
        row = worksheet[row_idx]
        cell_values = [str(cell.value or '').lower() for cell in row[:10]]  # Check first 10 columns
        matches = sum(1 for keyword in header_keywords if any(keyword in val for val in cell_values))
        
        if matches >= 2:  # At least 2 header keywords found
            return row_idx
    
    return 1  # Default to first row

def parse_xlsx_file(file_path: str, s3_key: Optional[str] = None) -> Dict[str, Any]:
    """Parse XLSX file and extract data with formulas"""
    # Load workbook
    if s3_key:
        # Download from S3
        file_data = download_from_s3(s3_key)
        workbook = openpyxl.load_workbook(file_data, data_only=False)  # data_only=False to preserve formulas
    else:
        workbook = openpyxl.load_workbook(file_path, data_only=False)
    
    result = {
        'sheets': [],
        'formulas_detected': 0,
        'volatile_formulas': 0,
        'total_rows': 0,
    }
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row = detect_header_row(worksheet)
        
        # Extract headers
        headers = []
        if header_row:
            header_row_cells = worksheet[header_row]
            headers = [str(cell.value or '').strip() for cell in header_row_cells[:20]]  # Max 20 columns
        
        # Extract data rows
        rows = []
        formulas_found = []
        
        for row_idx in range(header_row + 1 if header_row else 2, worksheet.max_row + 1):
            row_data = {}
            row_cells = worksheet[row_idx]
            
            # Skip empty rows
            if all(cell.value is None for cell in row_cells[:10]):
                continue
            
            for col_idx, cell in enumerate(row_cells[:len(headers)]):
                header = headers[col_idx] if col_idx < len(headers) else f'Column{col_idx + 1}'
                
                if cell.value is None:
                    continue
                
                # Check if cell has formula
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value)
                    formula_hash = compute_formula_hash(formula)
                    formula_refs = extract_formula_references(formula)
                    
                    # Get calculated value if available
                    calculated_value = None
                    try:
                        calculated_value = float(cell.value) if isinstance(cell.value, (int, float)) else None
                    except:
                        pass
                    
                    formulas_found.append({
                        'cell_ref': f"{sheet_name}!{cell.coordinate}",
                        'formula': formula,
                        'formula_hash': formula_hash,
                        'calculated_value': calculated_value,
                        'references': formula_refs,
                        'row': row_idx,
                        'col': col_idx + 1,
                    })
                    
                    result['formulas_detected'] += 1
                    if formula_refs['volatile']:
                        result['volatile_formulas'] += 1
                    
                    row_data[header] = {
                        'value': calculated_value,
                        'formula': formula,
                        'formula_hash': formula_hash,
                        'is_formula': True,
                        'cell_ref': f"{sheet_name}!{cell.coordinate}",
                    }
                else:
                    # Regular value
                    row_data[header] = cell.value
            
            if row_data:
                rows.append({
                    'row_index': row_idx,
                    'data': row_data,
                })
                result['total_rows'] += 1
        
        result['sheets'].append({
            'name': sheet_name,
            'header_row': header_row,
            'headers': headers,
            'rows': rows[:1000],  # Limit preview to 1000 rows
            'formulas': formulas_found,
            'total_rows': len(rows),
        })
    
    workbook.close()
    return result

def handle_xlsx_preview(job_id: str, org_id: str, object_id: str, logs: Dict[str, Any]):
    """Handle XLSX preview job - parse file and return preview"""
    try:
        # Extract params from logs
        params = logs.get('params', {}) if isinstance(logs, dict) else {}
        if not isinstance(params, dict):
            params = {}
        
        upload_key = params.get('uploadKey')
        s3_key = params.get('s3Key')
        file_hash = params.get('fileHash')
        
        logger.info(f"Starting XLSX preview for job {job_id}, org {org_id}")
        
        # Parse XLSX
        parsed_data = parse_xlsx_file('', s3_key=s3_key)
        
        # Store preview in job params or S3
        preview_data = {
            'sheets': parsed_data['sheets'],
            'formulas_detected': parsed_data['formulas_detected'],
            'volatile_formulas': parsed_data['volatile_formulas'],
            'total_rows': parsed_data['total_rows'],
            'file_hash': file_hash,
        }
        
        # Update job with preview
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE jobs
            SET status = 'completed',
                progress = 100,
                logs = jsonb_set(COALESCE(logs, '[]'::jsonb), '{-1}', %s::jsonb)
            WHERE id = %s
        """, (json.dumps({
            'level': 'info',
            'message': 'XLSX preview completed',
            'preview': preview_data,
            'ts': datetime.utcnow().isoformat(),
        }), job_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        logger.info(f"XLSX preview completed for job {job_id}")
        return preview_data
        
    except Exception as e:
        logger.error(f"Error in XLSX preview job {job_id}: {str(e)}", exc_info=True)
        raise

def handle_xlsx_import(job_id: str, org_id: str, object_id: str, logs: Dict[str, Any]):
    """Handle XLSX import job - create transactions with formula provenance"""
    conn = None
    cursor = None
    
    try:
        # Extract params from logs
        params = logs.get('params', {}) if isinstance(logs, dict) else {}
        if not isinstance(params, dict):
            params = {}
        
        upload_key = params.get('uploadKey')
        s3_key = params.get('s3Key')
        mapping_json = params.get('mappingJson', {})
        mapping_id = params.get('mappingId')
        
        logger.info(f"Starting XLSX import for job {job_id}, org {org_id}")
        
        # Parse XLSX
        parsed_data = parse_xlsx_file('', s3_key=s3_key)
        
        # Get connector for Excel
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id FROM connectors
            WHERE org_id = %s AND type = 'excel'
            LIMIT 1
        """, (org_id,))
        
        connector_result = cursor.fetchone()
        connector_id = connector_result[0] if connector_result else None
        
        if not connector_id:
            # Create Excel connector
            cursor.execute("""
                INSERT INTO connectors (org_id, type, status, created_at)
                VALUES (%s, 'excel', 'connected', NOW())
                RETURNING id
            """, (org_id,))
            connector_id = cursor.fetchone()[0]
        
        # Process each sheet
        transactions_created = 0
        formula_mappings = {}  # formula_hash -> assumption_id mapping
        
        for sheet_data in parsed_data['sheets']:
            sheet_name = sheet_data['name']
            headers = sheet_data['headers']
            rows = sheet_data['rows']
            formulas = sheet_data['formulas']
            
            # Apply column mappings
            column_mappings = mapping_json.get('columnMappings', {})
            
            # Prepare transactions for bulk insert
            transaction_values = []
            
            for row in rows:
                row_data = row['data']
                row_index = row['row_index']
                
                # Map columns
                date_value = None
                amount_value = None
                description_value = None
                category_value = None
                currency_value = 'USD'
                source_id_value = None
                raw_payload = {}
                
                # Apply mappings
                for target_field, mapping in column_mappings.items():
                    csv_column = mapping.get('csvColumn')
                    if csv_column and csv_column in row_data:
                        value = row_data[csv_column]
                        
                        if isinstance(value, dict) and value.get('is_formula'):
                            # Formula cell
                            raw_payload['formula'] = value.get('formula')
                            raw_payload['formula_hash'] = value.get('formula_hash')
                            raw_payload['cell_ref'] = value.get('cell_ref')
                            raw_payload['calculated_value'] = value.get('value')
                            
                            # Store formula reference
                            if value.get('formula_hash'):
                                formula_hash = value['formula_hash']
                                if formula_hash not in formula_mappings:
                                    formula_mappings[formula_hash] = {
                                        'formula': value.get('formula'),
                                        'cell_refs': [],
                                    }
                                formula_mappings[formula_hash]['cell_refs'].append(value.get('cell_ref'))
                            
                            value = value.get('value')  # Use calculated value
                        
                        if target_field == 'date':
                            date_value = value
                        elif target_field == 'amount':
                            amount_value = float(value) if value else 0
                        elif target_field == 'description':
                            description_value = str(value) if value else ''
                        elif target_field == 'category':
                            category_value = str(value) if value else None
                        elif target_field == 'currency':
                            currency_value = str(value) if value else 'USD'
                        elif target_field == 'source_id':
                            source_id_value = str(value) if value else None
                
                # Skip if no date or amount
                if not date_value or amount_value is None:
                    continue
                
                # Parse date
                try:
                    if isinstance(date_value, str):
                        # Try common date formats
                        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                            try:
                                date_value = datetime.strptime(date_value, fmt).date()
                                break
                            except:
                                continue
                    elif isinstance(date_value, datetime):
                        date_value = date_value.date()
                except:
                    logger.warning(f"Could not parse date: {date_value}")
                    continue
                
                # Prepare transaction
                transaction_values.append((
                    org_id,
                    connector_id,
                    source_id_value,
                    date_value,
                    amount_value,
                    currency_value,
                    category_value,
                    description_value,
                    json.dumps(raw_payload) if raw_payload else None,
                    datetime.utcnow(),
                ))
            
            # Bulk insert transactions
            if transaction_values:
                execute_values(
                    cursor,
                    """
                    INSERT INTO raw_transactions 
                    (org_id, connector_id, source_id, date, amount, currency, category, description, raw_payload, imported_at)
                    VALUES %s
                    """,
                    transaction_values,
                    template=None,
                    page_size=1000,
                )
                transactions_created += len(transaction_values)
        
        # Create Excel sync record
        cursor.execute("""
            INSERT INTO excel_syncs (org_id, file_hash, mapping_id, status, last_synced_at, metadata, created_at, updated_at)
            VALUES (%s, %s, %s, 'completed', NOW(), %s::jsonb, NOW(), NOW())
            ON CONFLICT DO NOTHING
        """, (org_id, params.get('fileHash'), mapping_id, json.dumps({
            'transactions_created': transactions_created,
            'formulas_detected': parsed_data['formulas_detected'],
            'volatile_formulas': parsed_data['volatile_formulas'],
            'formula_mappings': formula_mappings,
        })))
        
        conn.commit()
        
        logger.info(f"XLSX import completed: {transactions_created} transactions created")
        
        return {
            'transactions_created': transactions_created,
            'formulas_detected': parsed_data['formulas_detected'],
            'volatile_formulas': parsed_data['volatile_formulas'],
        }
        
    except Exception as e:
        logger.error(f"Error in XLSX import job {job_id}: {str(e)}", exc_info=True)
        if conn:
            conn.rollback()
        raise
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

